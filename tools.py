"""
GrimmBot — Built-in Tools Module
"""

import os
import io
import time
import re
import json
import difflib
import subprocess
import shutil
import hashlib
from datetime import datetime
from pathlib import Path
from typing import Optional

from config import AgentConfig, is_path_safe, is_domain_allowed, is_command_allowed
from memory import get_memory
from screen import (
    take_screenshot_raw, screenshot_to_base64,
    save_screenshot, mouse_move, mouse_click, mouse_double_click,
    mouse_scroll, mouse_drag, keyboard_type, keyboard_press,
    keyboard_shortcut, clipboard_copy, clipboard_paste, clipboard_get,
    clipboard_set, launch_chromium, close_chromium, chromium_navigate,
    chromium_new_tab, chromium_close_tab, chromium_switch_tab,
    chromium_refresh, chromium_back, chromium_forward, is_chromium_running,
    list_chromium_profiles, wipe_chromium_profile, get_active_window,
    focus_window, wait_for_screen_change, wait_for_screen_stable,
    read_true_dom, SCREEN_WIDTH, SCREEN_HEIGHT,
)

class Tools:
    def __init__(self, config: AgentConfig):
        self.config = config
        self.current_profile = "default"

    def get_current_time(self) -> str:
        return f"Current time: {datetime.now().isoformat()}"

    def wait(self, seconds: float = 2.0) -> str:
        time.sleep(min(seconds, 30))
        return f"Waited {seconds}s"

    def wait_for_change(self, timeout: float = 10.0) -> str:
        return wait_for_screen_change(timeout)

    def wait_for_stable(self, timeout: float = 10.0) -> str:
        return wait_for_screen_stable(timeout)

    def screenshot(self) -> str:
        return "Screenshot captured."

    def click(self, x: int, y: int, button: str = "left") -> str:
        return mouse_click(x, y, button)

    def double_click(self, x: int, y: int) -> str:
        return mouse_double_click(x, y)

    def move_mouse(self, x: int, y: int) -> str:
        return mouse_move(x, y)

    def drag(self, from_x: int, from_y: int, to_x: int, to_y: int) -> str:
        return mouse_drag(from_x, from_y, to_x, to_y)

    def scroll(self, direction: str = "down", amount: int = 3) -> str:
        return mouse_scroll(direction, amount)

    def type_text(self, text: str) -> str:
        return keyboard_type(text, human=True)

    def press_key(self, key: str) -> str:
        return keyboard_press(key)

    def hotkey(self, keys: str) -> str:
        return keyboard_shortcut(keys)

    def copy(self) -> str:
        return clipboard_copy()

    def paste(self) -> str:
        return clipboard_paste()

    def get_clipboard(self) -> str:
        return clipboard_get()

    def set_clipboard(self, text: str) -> str:
        return clipboard_set(text)

    def open_browser(self, url: str = "", profile: str = "") -> str:
        return launch_chromium(url, profile or self.current_profile)

    def close_browser(self) -> str:
        return close_chromium()

    def go_to_url(self, url: str) -> str:
        if not is_domain_allowed(url, self.config):
            return f"Domain not in allowlist: {url}"
        return chromium_navigate(url)

    def new_tab(self, url: str = "") -> str:
        if url and not is_domain_allowed(url, self.config):
            return f"Domain not in allowlist: {url}"
        return chromium_new_tab(url)

    def close_tab(self) -> str:
        return chromium_close_tab()

    def switch_tab(self, direction: str = "next") -> str:
        return chromium_switch_tab(direction)

    def refresh_page(self) -> str:
        return chromium_refresh()

    def go_back(self) -> str:
        return chromium_back()

    def go_forward(self) -> str:
        return chromium_forward()

    def read_dom(self) -> str:
        from screen import read_true_dom
        return read_true_dom()

    def click_element(self, element_id: int) -> str:
        from screen import INTERACTABLE_MAP, mouse_click
        element_id = str(element_id)
        if element_id not in INTERACTABLE_MAP:
            return f"Error: Element [ID: {element_id}] not found in current DOM view. Call read_dom() to refresh."
        coords = INTERACTABLE_MAP[element_id]
        return mouse_click(coords['x'], coords['y'], button="left")

    def get_active_window_title(self) -> str:
        return get_active_window()

    def focus_window_by_title(self, title: str) -> str:
        return focus_window(title)

    def read_file(self, path: str) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        if not p.exists():
            return f"Not found: {path}"
        if not p.is_file():
            return f"Not a file: {path}"
        if p.stat().st_size > 1_000_000:
            return "File too large (>1MB). Use shell."
        try:
            content = p.read_text(errors="replace")
            if len(content) > 20000:
                content = content[:20000] + "\n[...truncated...]"
            return f"=== {p.name} ===\n{content}"
        except Exception as e:
            return f"Read error: {e}"

    def read_file_lines(self, path: str, start_line: int = 1, end_line: int = 0) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        if not p.exists():
            return f"Not found: {path}"
        try:
            lines = p.read_text(errors="replace").splitlines()
            total = len(lines)
            s = max(1, start_line) - 1
            e = end_line if end_line > 0 else total
            e = min(e, total)
            numbered = [f"{i:4d} | {line}" for i, line in enumerate(lines[s:e], start=s+1)]
            return f"=== {p.name} (lines {s+1}-{e} of {total}) ===\n" + "\n".join(numbered)
        except Exception as e:
            return f"Read error: {e}"

    def write_file(self, path: str, content: str) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        wormhole = Path(self.config.wormhole_dir).resolve()
        if str(p.resolve()).startswith(str(wormhole)):
            for ext in self.config.wormhole_blocked_extensions:
                if p.name.lower().endswith(ext):
                    return f"Extension '{ext}' not allowed"
            if len(content.encode()) > self.config.wormhole_max_file_size:
                return "File too large"
        try:
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(content)
            return f"Written {len(content)} chars -> {p}"
        except Exception as e:
            return f"Write error: {e}"

    def patch_file(self, path: str, search: str, replace: str, occurrence: int = 1) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        if not p.exists():
            return f"Not found: {path}"
        try:
            content = p.read_text(errors="replace")
            count = content.count(search)
            if count == 0:
                close = difflib.get_close_matches(
                    search.splitlines()[0] if "\n" in search else search,
                    content.splitlines(), n=3, cutoff=0.4
                )
                hint = "\nSimilar:\n" + "\n".join(f"  -> {l}" for l in close) if close else ""
                return f"Pattern not found in {p.name}{hint}"
            if occurrence == 0:
                new = content.replace(search, replace)
                replaced = count
            else:
                if occurrence > count:
                    return f"Only {count} occurrences, requested #{occurrence}"
                parts = content.split(search)
                new = search.join(parts[:occurrence]) + replace + search.join(parts[occurrence:])
                replaced = 1
            p.write_text(new)
            diff = "\n".join(list(difflib.unified_diff(
                content.splitlines(keepends=True)[:50],
                new.splitlines(keepends=True)[:50],
                fromfile="before", tofile="after", lineterm=""
            ))[:30])
            return f"Patched {replaced} occurrence(s) in {p.name}\nDiff:\n{diff}"
        except Exception as e:
            return f"Patch error: {e}"

    def insert_at_line(self, path: str, line_number: int, content: str) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        if not p.exists():
            return f"Not found: {path}"
        try:
            lines = p.read_text(errors="replace").splitlines(keepends=True)
            idx = max(0, min(line_number - 1, len(lines)))
            insert = [l if l.endswith("\n") else l + "\n" for l in content.splitlines()]
            lines[idx:idx] = insert
            p.write_text("".join(lines))
            return f"Inserted {len(insert)} lines at line {line_number} in {p.name}"
        except Exception as e:
            return f"Insert error: {e}"

    def delete_lines(self, path: str, start_line: int, end_line: int) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        if not p.exists():
            return f"Not found: {path}"
        try:
            lines = p.read_text(errors="replace").splitlines(keepends=True)
            s, e = max(0, start_line - 1), min(end_line, len(lines))
            deleted = lines[s:e]
            del lines[s:e]
            p.write_text("".join(lines))
            preview = "".join(deleted[:5])
            if len(deleted) > 5:
                preview += f"...({len(deleted)-5} more)"
            return f"Deleted lines {start_line}-{end_line} from {p.name}:\n{preview}"
        except Exception as e:
            return f"Delete error: {e}"

    def list_directory(self, path: str) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        if not p.exists():
            return f"Not found: {path}"
        try:
            lines = [f"{path}:"]
            for item in sorted(p.iterdir())[:100]:
                if item.is_dir():
                    lines.append(f"  [dir] {item.name}/")
                else:
                    lines.append(f"  {item.name} ({item.stat().st_size:,} bytes)")
            return "\n".join(lines)
        except Exception as e:
            return f"Error: {e}"

    def delete_file(self, path: str) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        p = Path(path)
        if not p.exists():
            return f"Not found: {path}"
        try:
            shutil.rmtree(p) if p.is_dir() else p.unlink()
            return f"Deleted: {path}"
        except Exception as e:
            return f"Error: {e}"

    def find_in_files(self, directory: str, pattern: str, file_glob: str = "*") -> str:
        if not is_path_safe(directory):
            return "Path outside allowed directories"
        p = Path(directory)
        if not p.exists():
            return f"Not found: {directory}"
        try:
            results = []
            for fp in sorted(p.rglob(file_glob))[:200]:
                if not fp.is_file() or fp.stat().st_size > 500_000:
                    continue
                try:
                    for i, line in enumerate(fp.read_text(errors="replace").splitlines(), 1):
                        if pattern.lower() in line.lower():
                            results.append(f"  {fp.relative_to(p)}:{i}: {line.strip()[:100]}")
                            if len(results) >= 50:
                                return f"Found '{pattern}':\n" + "\n".join(results + ["  ...(truncated)"])
                except Exception:
                    continue
            if not results:
                return f"'{pattern}' not found in {directory}"
            return f"Found '{pattern}':\n" + "\n".join(results)
        except Exception as e:
            return f"Search error: {e}"

    def text_to_pdf(self, input_path: str, output_path: str) -> str:
        if not is_path_safe(input_path) or not is_path_safe(output_path):
            return "Path outside allowed directories"
        inp = Path(input_path)
        if not inp.exists():
            return f"Not found: {input_path}"
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            text = inp.read_text(errors="replace")
            outp = Path(output_path)
            outp.parent.mkdir(parents=True, exist_ok=True)
            c = canvas.Canvas(str(outp), pagesize=letter)
            w, h = letter
            y = h - 50
            for line in text.splitlines():
                if y < 50:
                    c.showPage()
                    y = h - 50
                while len(line) > 90:
                    c.drawString(50, y, line[:90])
                    line = line[90:]
                    y -= 14
                    if y < 50:
                        c.showPage()
                        y = h - 50
                c.drawString(50, y, line)
                y -= 14
            c.save()
            return f"PDF created: {outp}"
        except Exception as e:
            return f"PDF error: {e}"

    def convert_document(self, input_path: str, output_path: str) -> str:
        if not is_path_safe(input_path) or not is_path_safe(output_path):
            return "Path outside allowed directories"
        try:
            outp = Path(output_path)
            outp.parent.mkdir(parents=True, exist_ok=True)
            result = subprocess.run(["pandoc", input_path, "-o", str(outp)],
                                    capture_output=True, text=True, timeout=60)
            return f"Converted: {outp}" if result.returncode == 0 else f"Pandoc error: {result.stderr}"
        except Exception as e:
            return f"Error: {e}"

    def read_excel(self, path: str, sheet: str = "") -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            target = sheet or wb.sheetnames[0]
            if target not in wb.sheetnames:
                return f"Sheet '{target}' not found. Available: {wb.sheetnames}"
            lines = []
            for i, row in enumerate(wb[target].iter_rows(values_only=True)):
                if i > 500:
                    lines.append("[...truncated...]")
                    break
                lines.append(" | ".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(lines)
        except Exception as e:
            return f"Excel error: {e}"

    def write_excel(self, path: str, data: list) -> str:
        if not is_path_safe(path):
            return "Path outside allowed directories"
        try:
            from openpyxl import Workbook
            outp = Path(path)
            wormhole = Path(self.config.wormhole_dir).resolve()
            outp.parent.mkdir(parents=True, exist_ok=True)
            wb = Workbook()
            ws = wb.active
            for row in data:
                ws.append(row if isinstance(row, list) else [row])
            wb.save(str(outp))
            return f"Written {len(data)} rows -> {outp}"
        except Exception as e:
            return f"Error: {e}"

    def shell(self, command: str, cwd: str = "") -> str:
        if not is_command_allowed(command, self.config):
            return "Command not in allowlist or contains blocked characters"
        work_dir = cwd or self.config.workspace_dir
        if cwd and not is_path_safe(cwd):
            return "Directory outside allowed paths"
        try:
            Path(work_dir).mkdir(parents=True, exist_ok=True)
            result = subprocess.run(
                command, shell=True, cwd=work_dir,
                capture_output=True, text=True, timeout=120,
                env={**os.environ, "HOME": "/home/grimmbot"},
            )
            output = f"EXIT: {result.returncode}\n"
            if result.stdout:
                output += f"STDOUT:\n{result.stdout[:self.config.max_shell_output]}\n"
            if result.stderr:
                output += f"STDERR:\n{result.stderr[:self.config.max_shell_output]}\n"
            if not result.stdout and not result.stderr:
                output += "(no output)\n"
            return output
        except subprocess.TimeoutExpired:
            return "Command timed out (120s)"
        except Exception as e:
            return f"Shell error: {e}"

    def remember(self, information: str, tags: list = None) -> str:
        try:
            get_memory(self.current_profile).add(task="Note", result=information, tags=tags or ["note"])
            return f"Remembered: {information[:100]}"
        except Exception as e:
            return f"Error: {e}"

    def recall(self, query: str) -> str:
        try:
            results = get_memory(self.current_profile).search(query, top_k=5)
            if not results:
                return "No relevant memories found."
            return "Found:\n" + "\n".join(f"  [{e.timestamp}] {e.result_summary[:100]}" for e in results)
        except Exception as e:
            return f"Error: {e}"

    def schedule_task(self, prompt: str, time_str: str) -> str:
        try:
            from scheduler import get_scheduler
            sched = get_scheduler()
            now = datetime.now(sched.tz)
            if len(time_str) == 5 and ":" in time_str:
                h, m = map(int, time_str.split(":"))
                target = now.replace(hour=h, minute=m, second=0, microsecond=0)
                if target <= now:
                    target += __import__("datetime").timedelta(days=1)
            else:
                target = datetime.fromisoformat(time_str)
                if target.tzinfo is None:
                    target = sched.tz.localize(target)
            return f"Scheduled for {target.strftime('%Y-%m-%d %H:%M')}: {sched.schedule_once(prompt, target, self.current_profile)}"
        except Exception as e:
            return f"Error: {e}"

    def schedule_daily(self, prompt: str, time_str: str) -> str:
        try:
            from scheduler import get_scheduler
            h, m = map(int, time_str.split(":"))
            return f"Daily at {time_str}: {get_scheduler().schedule_daily(prompt, h, m, self.current_profile)}"
        except Exception as e:
            return f"Error: {e}"

    def schedule_interval(self, prompt: str, minutes: int) -> str:
        try:
            from scheduler import get_scheduler
            return f"Every {minutes}min: {get_scheduler().schedule_interval(prompt, minutes, self.current_profile)}"
        except Exception as e:
            return f"Error: {e}"

    def list_scheduled_tasks(self) -> str:
        try:
            from scheduler import get_scheduler
            tasks = get_scheduler().list_tasks()
            if not tasks:
                return "No scheduled tasks."
            return "Scheduled:\n" + "\n".join(
                f"  {'[ON]' if t.enabled else '[OFF]'} {t.id}: {t.prompt[:40]}..." for t in tasks
            )
        except Exception as e:
            return f"Error: {e}"

    def cancel_scheduled_task(self, task_id: str) -> str:
        try:
            from scheduler import get_scheduler
            return f"Cancelled: {task_id}" if get_scheduler().cancel(task_id) else f"Not found: {task_id}"
        except Exception as e:
            return f"Error: {e}"

    def list_profiles(self) -> str:
        return list_chromium_profiles()

    def wipe_profile(self, profile: str) -> str:
        return wipe_chromium_profile(profile)

    def done(self, result: str) -> str:
        return result

    def create_plan(self, goal: str, steps: list) -> str:
        plan_file = Path(self.config.workspace_dir) / "current_plan.md"
        plan_file.parent.mkdir(parents=True, exist_ok=True)
        lines = ["# Execution Plan", f"**Goal:** {goal}",
                 f"**Created:** {datetime.now().isoformat()}", "**Status:** In Progress", "", "## Steps"]
        for i, step in enumerate(steps, 1):
            lines.append(f"{i}. [ ] {step}")
        plan_file.write_text("\n".join(lines))
        return f"Plan created with {len(steps)} steps -> {plan_file}"

    def update_plan_step(self, step_number: int, status: str = "done", notes: str = "") -> str:
        plan_file = Path(self.config.workspace_dir) / "current_plan.md"
        if not plan_file.exists():
            return "No active plan found"
        try:
            content = plan_file.read_text()
            markers = {"done": "[x]", "failed": "[!]", "skipped": "[-]"}
            marker = markers.get(status, "[x]")
            lines = content.splitlines()
            pat = re.compile(r"(\d+)\. \[[ x!\-]\] (.+)")
            for i, line in enumerate(lines):
                m = pat.match(line)
                if m and int(m.group(1)) == step_number:
                    note = f" -- {notes}" if notes else ""
                    lines[i] = f"{step_number}. {marker} {m.group(2)}{note}"
                    plan_file.write_text("\n".join(lines))
                    return f"Step {step_number} marked as {status}"
            return f"Step {step_number} not found"
        except Exception as e:
            return f"Error: {e}"

    def create_custom_tool(self, name: str, description: str, parameters: dict, code: str) -> str:
        return "Routed internally"

    def list_custom_tools(self) -> str:
        return "Routed internally"

    def delete_custom_tool(self, name: str) -> str:
        return "Routed internally"

    def save_adaptation_rule(self, rule: str) -> str:
        """Allows the agent to actively save its own rules."""
        adap_path = Path(self.config.adaptation_file)
        adaptations = []
        
        if adap_path.exists():
            try:
                adaptations = json.loads(adap_path.read_text())
            except Exception:
                pass
        
        if rule not in adaptations:
            adaptations.append(rule)
            adap_path.write_text(json.dumps(adaptations, indent=2))
            return f"SUCCESS: Rule '{rule}' saved to adaptation.json. You will remember this for future tasks."
            
        return "Rule already exists in adaptation.json."